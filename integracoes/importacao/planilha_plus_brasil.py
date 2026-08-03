"""
integracoes/importacao/planilha_plus_brasil.py
Parser da planilha PLUS BRASIL (custos na importação) + ponte para custo_landed.

A planilha é fonte de inputs e checklist de despesas — a cascata tributária
atualizada permanece em custo_landed (AFRMM 8%, Siscomex Portaria ME, etc.).
"""
from __future__ import annotations

import json
import logging
import re
from pathlib import Path
from typing import Any

from core.config import ROOT
from core.datadog_metrics import gauge, incrementar

logger = logging.getLogger("planilha_plus_brasil")

PLANILHA_DEFAULT = ROOT / "dados" / "importacao_simula_plus_brasil.xlsx"
CATALOGO_DESPESAS = ROOT / "catalogo" / "importacao_despesas_plus.json"

# Rótulo na planilha → id estável
_MAP_DESPESAS: list[tuple[str, str]] = [
    ("liberacao de conhecimento", "liberacao_conhecimento"),
    ("desconsolidacao de conhecimento", "desconsolidacao"),
    ("desconsolidação de conhecimento", "desconsolidacao"),
    ("marinha mercante", "afrmm"),
    ("armazenagem na zona primaria", "armazenagem_zp"),
    ("armazenagem na zona primária", "armazenagem_zp"),
    ("capatazias", "capatazias"),
    ("transporte para  importador", "transporte_importador"),
    ("transporte para importador", "transporte_importador"),
    ("s.d.a.", "sda_desembaraco"),
    ("desembaraco/logistica", "desembaraco_logistica"),
    ("desembaraço/logística", "desembaraco_logistica"),
    ("fumigacao", "fumigacao"),
    ("fumigação", "fumigacao"),
    ("licenciamento de importacao", "licenciamento"),
    ("licenciamento de importação", "licenciamento"),
    ("rpa", "rpa"),
]


def _norm(texto: Any) -> str:
    t = str(texto or "").strip().lower()
    t = (
        t.replace("ç", "c")
        .replace("á", "a")
        .replace("à", "a")
        .replace("ã", "a")
        .replace("â", "a")
        .replace("é", "e")
        .replace("ê", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ô", "o")
        .replace("õ", "o")
        .replace("ú", "u")
    )
    return re.sub(r"\s+", " ", t)


def _f(v: Any, default: float = 0.0) -> float:
    try:
        if v is None or v == "":
            return default
        return float(v)
    except (TypeError, ValueError):
        return default


def _pct_para_percentual(v: Any) -> float:
    """Planilha guarda 0.14 para 14% — converte para escala 0–100 do motor."""
    x = _f(v)
    if 0 < x <= 1.0:
        return round(x * 100.0, 4)
    return x


def caminho_planilha_plus(path: str | Path | None = None) -> Path:
    if path:
        return Path(path)
    try:
        from core.config import IMPORTACAO_PLANILHA_PLUS

        cfg = str(IMPORTACAO_PLANILHA_PLUS or "").strip()
        if cfg:
            p = Path(cfg)
            return p if p.is_absolute() else ROOT / p
    except Exception:
        pass
    return PLANILHA_DEFAULT


def carregar_despesas_padrao() -> dict[str, float]:
    try:
        data = json.loads(CATALOGO_DESPESAS.read_text(encoding="utf-8"))
        raw = data.get("despesas_padrao_brl") or {}
        return {str(k): _f(v) for k, v in raw.items()}
    except (OSError, json.JSONDecodeError, TypeError):
        return {}


def _ler_matriz(path: Path) -> list[list[Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb.active
        return [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _cel(matriz: list[list[Any]], r: int, c: int) -> Any:
    if r < 0 or r >= len(matriz):
        return None
    row = matriz[r]
    if c < 0 or c >= len(row):
        return None
    return row[c]


def _id_despesa(rotulo: str) -> str | None:
    n = _norm(rotulo)
    for trecho, did in _MAP_DESPESAS:
        if _norm(trecho) in n:
            return did
    return None


def parsear_planilha_plus(path: str | Path | None = None) -> dict[str, Any]:
    """Lê layout PLUS BRASIL → struct tipado (inputs + totais da planilha)."""
    p = caminho_planilha_plus(path)
    if not p.exists():
        return {"ok": False, "motivo": f"planilha não encontrada: {p}"}

    try:
        matriz = _ler_matriz(p)
    except Exception as exc:
        return {"ok": False, "motivo": f"falha ao ler planilha: {exc}"}

    cambio = _f(_cel(matriz, 10, 1))  # Dólar linha típica
    # Algumas cópias repetem câmbio na linha seguinte
    if cambio <= 0:
        cambio = _f(_cel(matriz, 11, 1))

    vmle_usd = _f(_cel(matriz, 15, 3))
    vmle_brl = _f(_cel(matriz, 15, 5))
    frete_usd = _f(_cel(matriz, 16, 3))
    frete_brl = _f(_cel(matriz, 16, 5))
    seguro_brl = _f(_cel(matriz, 17, 5))
    vmld_brl = _f(_cel(matriz, 18, 5))
    acrescimos_brl = _f(_cel(matriz, 19, 5))
    cif_brl = _f(_cel(matriz, 20, 5))

    ncm = str(_cel(matriz, 23, 0) or "").strip()

    ii_pct = _pct_para_percentual(_cel(matriz, 25, 5) or _cel(matriz, 25, 1))
    ipi_pct = _pct_para_percentual(_cel(matriz, 26, 5) or _cel(matriz, 26, 1))
    pis_pct = _pct_para_percentual(_cel(matriz, 27, 5) or _cel(matriz, 27, 1))
    cofins_pct = _pct_para_percentual(_cel(matriz, 28, 5) or _cel(matriz, 28, 1))
    icms_pct = _pct_para_percentual(_cel(matriz, 29, 5) or _cel(matriz, 29, 1))

    siscomex_planilha = _f(_cel(matriz, 34, 4))
    afrmm_pct_planilha = _pct_para_percentual(_cel(matriz, 59, 3))
    afrmm_brl_planilha = _f(_cel(matriz, 59, 4))

    outras: dict[str, float] = {}
    outras_linhas: list[dict[str, Any]] = []
    for r in range(55, 70):
        rotulo = _cel(matriz, r, 0)
        if not rotulo:
            continue
        n = _norm(rotulo)
        if "total das despesas" in n:
            break
        did = _id_despesa(str(rotulo))
        if not did:
            continue
        valor = _f(_cel(matriz, r, 4))
        outras[did] = valor
        outras_linhas.append({"id": did, "rotulo": str(rotulo).strip(), "valor_brl": valor})

    total_despesas_planilha = _f(_cel(matriz, 68, 4))
    impostos_planilha = _f(_cel(matriz, 74, 4))
    if impostos_planilha <= 0:
        impostos_planilha = _f(_cel(matriz, 53, 1))
    total_geral_planilha = _f(_cel(matriz, 76, 4))
    mercadoria_brl = _f(_cel(matriz, 72, 4))
    frete_seguro_brl = _f(_cel(matriz, 73, 4))

    ok = cambio > 0 and (vmle_usd > 0 or vmle_brl > 0) and cif_brl > 0
    out = {
        "ok": ok,
        "fonte": str(p),
        "layout": "plus_brasil",
        "cambio_usd_brl": round(cambio, 4),
        "vmle_usd": round(vmle_usd, 4),
        "vmle_brl": round(vmle_brl, 2),
        "frete_internacional_usd": round(frete_usd, 4),
        "frete_internacional_brl": round(frete_brl, 2),
        "seguro_brl": round(seguro_brl, 2),
        "vmld_brl": round(vmld_brl, 2),
        "acrescimos_brl": round(acrescimos_brl, 2),
        "cif_brl": round(cif_brl, 2),
        "ncm": ncm,
        "aliquotas": {
            "ii_pct": ii_pct,
            "ipi_pct": ipi_pct,
            "pis_pct": pis_pct,
            "cofins_pct": cofins_pct,
            "icms_pct": icms_pct,
        },
        "siscomex_planilha_brl": round(siscomex_planilha, 2),
        "afrmm_pct_planilha": afrmm_pct_planilha,
        "afrmm_brl_planilha": round(afrmm_brl_planilha, 2),
        "outras_despesas": outras,
        "outras_despesas_linhas": outras_linhas,
        "totais_planilha": {
            "mercadoria_brl": round(mercadoria_brl, 2),
            "frete_seguro_brl": round(frete_seguro_brl, 2),
            "impostos_brl": round(impostos_planilha, 2),
            "outras_despesas_brl": round(total_despesas_planilha, 2),
            "total_geral_brl": round(total_geral_planilha, 2),
        },
        "avisos": [
            "AFRMM da planilha (ex.: 25%) é legado — motor usa 8% (Lei 14.301/2022).",
            "Siscomex 214,50 da planilha é legado — motor usa Portaria ME vigente.",
            "Admissão temporária da planilha é ignorada (fluxo marketplace).",
        ],
    }
    if not ok:
        out["motivo"] = "campos essenciais ausentes (câmbio/VMLE/CIF)"
    return out


def _despesas_para_motor(parsed: dict[str, Any]) -> tuple[dict[str, float], float, float]:
    """
    Separa despesas da planilha:
      - outras_despesas (base ICMS, sem AFRMM/siscomex)
      - desembaraco_brl (SDA + desembaraço/logística)
      - frete_nacional_total (transporte importador)
    """
    raw = dict(parsed.get("outras_despesas") or {})
    transporte = _f(raw.pop("transporte_importador", 0))
    raw.pop("afrmm", None)
    sda = _f(raw.pop("sda_desembaraco", 0))
    desp_log = _f(raw.pop("desembaraco_logistica", 0))
    desembaraco = sda + desp_log
    # remove zeros
    outras = {k: round(v, 2) for k, v in raw.items() if _f(v) > 0}
    return outras, desembaraco, transporte


def calcular_desde_planilha_plus(
    path: str | Path | None = None,
    *,
    usar_despesas_planilha: bool = True,
    modo_frete: str = "maritimo",
) -> dict[str, Any]:
    """
    Parse PLUS → custo_landed com cascata vigente.
    Retorna planilha_totais vs motor_totais + deltas conhecidos.
    """
    from integracoes.importacao.custo_landed import calcular_custo_landed
    from integracoes.importacao.siscomex import taxa_siscomex_brl

    parsed = parsear_planilha_plus(path)
    if not parsed.get("ok"):
        incrementar("importacao.plus_brasil_erro")
        return parsed

    cambio = _f(parsed.get("cambio_usd_brl"))
    vmle_usd = _f(parsed.get("vmle_usd"))
    if vmle_usd <= 0 and cambio > 0:
        vmle_usd = _f(parsed.get("vmle_brl")) / cambio

    aliq = parsed.get("aliquotas") or {}
    frete_usd = _f(parsed.get("frete_internacional_usd"))
    seguro = _f(parsed.get("seguro_brl"))
    acrescimos = _f(parsed.get("acrescimos_brl"))

    if usar_despesas_planilha:
        outras, desembaraco, frete_nac_total = _despesas_para_motor(parsed)
    else:
        padrao = carregar_despesas_padrao()
        outras, desembaraco, frete_nac_total = _despesas_para_motor(
            {"outras_despesas": padrao}
        )

    # Lote = 1 unidade “embarque” (planilha é por DI, não por SKU unitário)
    qty = 1
    frete_nacional_unit = frete_nac_total  # qty=1

    motor = calcular_custo_landed(
        vmle_usd,
        cambio_usd_brl=cambio,
        peso_kg_unit=1.0,
        quantidade=qty,
        modo_frete=modo_frete,  # type: ignore[arg-type]
        ii_pct=_f(aliq.get("ii_pct"), 16.0),
        ipi_pct=_f(aliq.get("ipi_pct"), 0.0),
        pis_pct=_f(aliq.get("pis_pct"), 2.1),
        cofins_pct=_f(aliq.get("cofins_pct"), 9.65),
        icms_pct=_f(aliq.get("icms_pct"), 18.0),
        seguro_pct=0.0,
        seguro_brl=seguro,
        frete_internacional_usd=frete_usd,
        acrescimos_brl=acrescimos,
        desembaraco_brl=desembaraco if desembaraco > 0 else 800.0,
        frete_nacional_brl_unit=frete_nacional_unit,
        outras_despesas_brl=outras,
        siscomex_brl=None,  # força regra vigente
        afrmm_pct=None,  # força 8% marítimo
    )

    totais_p = parsed.get("totais_planilha") or {}
    siscomex_vigente = taxa_siscomex_brl(adicoes=1)
    deltas = {
        "afrmm_pct_planilha": _f(parsed.get("afrmm_pct_planilha")),
        "afrmm_pct_motor": _f(motor.get("afrmm_pct")),
        "afrmm_brl_planilha": _f(parsed.get("afrmm_brl_planilha")),
        "afrmm_brl_motor": _f(motor.get("afrmm_brl")),
        "siscomex_brl_planilha": _f(parsed.get("siscomex_planilha_brl")),
        "siscomex_brl_motor": _f(motor.get("siscomex_brl"), siscomex_vigente),
        "total_geral_planilha": _f(totais_p.get("total_geral_brl")),
        "custo_total_motor": _f(motor.get("custo_total_brl")),
        "delta_total_brl": round(
            _f(motor.get("custo_total_brl")) - _f(totais_p.get("total_geral_brl")), 2
        ),
    }

    out = {
        "ok": bool(motor.get("ok")),
        "fonte": parsed.get("fonte"),
        "ncm": parsed.get("ncm"),
        "planilha": parsed,
        "motor": motor,
        "planilha_totais": totais_p,
        "motor_totais": {
            "cif_brl": motor.get("cif_brl"),
            "impostos_total_brl": motor.get("impostos_total_brl"),
            "despesas_aduaneiras_brl": motor.get("despesas_aduaneiras_brl"),
            "frete_nacional_brl": motor.get("frete_nacional_brl"),
            "custo_total_brl": motor.get("custo_total_brl"),
            "custo_unitario_brl": motor.get("custo_unitario_brl"),
        },
        "deltas": deltas,
        "avisos": parsed.get("avisos") or [],
    }

    tags = ["fonte:plus_brasil", f"modal:{modo_frete}"]
    incrementar("importacao.plus_brasil_ok", tags=tags)
    gauge("importacao.plus_brasil_total_motor", _f(motor.get("custo_total_brl")), tags)
    gauge("importacao.plus_brasil_delta_total", _f(deltas.get("delta_total_brl")), tags)
    logger.info(
        "plus_brasil: ncm=%s cif_planilha=%.2f custo_motor=%.2f delta_total=%.2f "
        "afrmm_planilha_pct=%.1f afrmm_motor_pct=%.1f siscomex_p=%.2f siscomex_m=%.2f",
        parsed.get("ncm"),
        _f(parsed.get("cif_brl")),
        _f(motor.get("custo_total_brl")),
        _f(deltas.get("delta_total_brl")),
        _f(deltas.get("afrmm_pct_planilha")),
        _f(deltas.get("afrmm_pct_motor")),
        _f(deltas.get("siscomex_brl_planilha")),
        _f(deltas.get("siscomex_brl_motor")),
    )
    return out
