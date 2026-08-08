# -*- coding: utf-8 -*-
"""
integracoes/esmaltes/planilha_consolidado_ecommerce.py

Lê planilhas_ecommerce/Consolidado (+ Custo Embalagem) e materializa:
  - plano de validação (invest + meta avaliações)
  - kits Impala do Resumo (custo/frete/fases)
  - kits Cruzeiro + ranking margens ML/Shopee
  - oportunidades Impala (POV / Ju Paes / complementos)
  - complementos Lívia
  - faixas de frete por peso

Também atualiza frete_estimado em catalogo/produtos.json e mescla kits
de validação que ainda não existem no catálogo operacional.
"""
from __future__ import annotations

import logging
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from core.atomic_io import escrever_json_atomico, ler_json
from core.config import ROOT
from core.datadog_metrics import gauge, incrementar

logger = logging.getLogger("planilha_consolidado_ecommerce")

CONSOLIDADO_DEFAULT = ROOT / "planilhas_ecommerce" / "Consolidado_Impala_Cruzeiro.xlsx"
EMBALAGEM_DEFAULT = (
    ROOT / "planilhas_ecommerce" / "planinhas" / "Custo_Embalagem_Kits_Impala.xlsx"
)
INTEL_DEFAULT = (
    ROOT / "planilhas_ecommerce" / "planinhas" / "Kit_Inteligencia_Beleza_Impala_v3.xlsx"
)

CAT_PLANO = ROOT / "catalogo" / "plano_validacao_impala.json"
CAT_CRUZEIRO = ROOT / "catalogo" / "kits_cruzeiro.json"
CAT_OPORT = ROOT / "catalogo" / "oportunidades_impala.json"
CAT_LIVIA = ROOT / "catalogo" / "complementos_livia.json"
CAT_FRETE = ROOT / "catalogo" / "frete_faixas_impala.json"
CAT_PRODUTOS = ROOT / "catalogo" / "produtos.json"

_SKU_RE = re.compile(r"\b((?:IMP|CRZ|BUNDLE)-[A-Z0-9]+(?:-[A-Z0-9]+)*)\b", re.I)
_SKU_OK = (
    re.compile(r"^IMP-[A-Z]+-\d+[A-Z]?$", re.I),
    re.compile(r"^CRZ-KIT-\d+$", re.I),
    re.compile(r"^BUNDLE-[A-Z0-9-]+$", re.I),
)
_MONEY_RE = re.compile(r"[-+]?\d+[.,]?\d*")


def _f(val: Any, default: float = 0.0) -> float:
    if val is None or val == "":
        return default
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s:
        return default
    # "R$ 10.48 (23%)" / "0.519" / "23%"
    if "%" in s and "R$" not in s.upper() and "(" not in s:
        m = _MONEY_RE.search(s.replace(",", "."))
        return float(m.group(0)) if m else default
    # dinheiro BR às vezes usa vírgula decimal
    m = _MONEY_RE.search(s.replace("R$", "").replace(" ", ""))
    if not m:
        return default
    raw = m.group(0)
    if "," in raw and "." in raw:
        raw = raw.replace(".", "").replace(",", ".")
    elif "," in raw:
        raw = raw.replace(",", ".")
    try:
        return float(raw)
    except ValueError:
        return default


def _pct_lucro_cell(val: Any) -> float | None:
    """Extrai % de 'R$ 10.48 (23%)' ou fração 0.23."""
    if val is None:
        return None
    s = str(val)
    m = re.search(r"\(([-+]?\d+[.,]?\d*)\s*%\)", s)
    if m:
        return _f(m.group(1))
    if isinstance(val, float) and 0 < abs(val) <= 1.5:
        return round(val * 100.0, 2)
    return None


def _sku_de(cell: Any) -> str:
    s = str(cell or "").strip().upper()
    m = _SKU_RE.search(s)
    if not m:
        return ""
    sku = m.group(1).upper()
    if not any(p.match(sku) for p in _SKU_OK):
        # tenta só o primeiro token estilo IMP-XXX-123
        m2 = re.match(r"^((?:IMP|CRZ|BUNDLE)-[A-Z0-9]+(?:-[A-Z0-9]+)?)", sku)
        if m2 and any(p.match(m2.group(1)) for p in _SKU_OK):
            return m2.group(1)
        return ""
    return sku


def _nome_apos_sku(cell: Any, sku: str) -> str:
    s = str(cell or "").strip()
    if not sku:
        return s
    # remove SKU no início
    resto = re.sub(rf"^{re.escape(sku)}\s*", "", s, flags=re.I).strip(" -—")
    return resto or s


def _sheet_rows(path: Path, sheet: str) -> list[tuple[Any, ...]]:
    import openpyxl

    if not path.is_file():
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        return []
    rows = [tuple(r) for r in wb[sheet].iter_rows(values_only=True)]
    wb.close()
    return rows


def _achar_header(rows: list[tuple[Any, ...]], *must_contain: str) -> int | None:
    needles = [n.lower() for n in must_contain]
    for i, row in enumerate(rows[:40]):
        blob = " ".join(str(c or "").lower().replace("\n", " ") for c in row[:12])
        if all(n in blob for n in needles):
            return i
    return None


def frete_para_peso_gramas(
    peso_gramas: float,
    faixas: list[dict[str, Any]] | None = None,
) -> float:
    """Retorna frete R$ pela menor faixa cujo teto_g >= peso."""
    bands = faixas if faixas is not None else carregar_frete_faixas()
    if not bands:
        return 0.0
    peso = max(0.0, float(peso_gramas or 0))
    ordenadas = sorted(bands, key=lambda b: float(b.get("teto_g") or 0))
    for b in ordenadas:
        if peso <= float(b.get("teto_g") or 0):
            return float(b.get("frete_reais") or 0)
    return float(ordenadas[-1].get("frete_reais") or 0)


def parse_frete_faixas(path: Path | None = None) -> list[dict[str, Any]]:
    p = path or EMBALAGEM_DEFAULT
    if not p.is_file():
        p = CONSOLIDADO_DEFAULT
    rows = _sheet_rows(p, "Custo por Kit")
    out: list[dict[str, Any]] = []
    for row in rows:
        item = str(row[0] or "")
        if "frete estimado" not in item.lower():
            continue
        m = re.search(r"([\d.,]+)\s*kg|([\d.,]+)\s*g", item, re.I)
        teto_g = 0.0
        if m:
            if m.group(1):
                teto_g = _f(m.group(1)) * 1000.0
            else:
                teto_g = _f(m.group(2))
        # "≤350g" / "≤1,0kg"
        m2 = re.search(r"[≤<=]\s*([\d.,]+)\s*(kg|g)", item, re.I)
        if m2:
            v = _f(m2.group(1))
            teto_g = v * 1000.0 if m2.group(2).lower() == "kg" else v
        frete = _f(row[1] if len(row) > 1 else 0)
        if teto_g > 0 and frete > 0:
            out.append(
                {
                    "teto_g": teto_g,
                    "frete_reais": frete,
                    "rotulo": item.strip(),
                }
            )
    out.sort(key=lambda x: x["teto_g"])
    return out


def carregar_frete_faixas() -> list[dict[str, Any]]:
    data = ler_json(CAT_FRETE, default={})
    if isinstance(data, dict) and isinstance(data.get("faixas"), list) and data["faixas"]:
        return data["faixas"]
    return parse_frete_faixas()


def parse_plano_validacao(path: Path | None = None) -> list[dict[str, Any]]:
    rows = _sheet_rows(path or CONSOLIDADO_DEFAULT, "Plano de Validação ML")
    hdr = _achar_header(rows, "invest", "meta")
    if hdr is None:
        hdr = _achar_header(rows, "kit", "valida")
    if hdr is None:
        return []
    out: list[dict[str, Any]] = []
    for row in rows[hdr + 1 :]:
        sku = _sku_de(row[0] if row else "")
        if not sku.startswith("IMP-"):
            continue
        meta_txt = str(row[6] if len(row) > 6 else "")
        m_meta = re.search(r"(\d+)", meta_txt)
        out.append(
            {
                "sku": sku,
                "nome": _nome_apos_sku(row[0], sku),
                "valida_unidades": int(_f(row[1] if len(row) > 1 else 0)),
                "custo_total_kit": round(_f(row[2] if len(row) > 2 else 0), 2),
                "preco_fase1": round(_f(row[3] if len(row) > 3 else 0), 2),
                "lucro_fase1_reais": round(_f(row[4] if len(row) > 4 else 0), 2),
                "lucro_fase1_pct": _pct_lucro_cell(row[4] if len(row) > 4 else None),
                "invest_validacao_reais": round(_f(row[5] if len(row) > 5 else 0), 2),
                "meta_avaliacoes": int(m_meta.group(1)) if m_meta else 20,
                "meta_avaliacoes_txt": meta_txt.strip(),
                "acao_pos_validar": str(row[7] if len(row) > 7 else "").strip(),
                "fonte": "Consolidado_Impala_Cruzeiro.xlsx / Plano de Validação ML",
            }
        )
    return out


def parse_resumo_kits(path: Path | None = None) -> list[dict[str, Any]]:
    rows = _sheet_rows(path or CONSOLIDADO_DEFAULT, "Resumo dos Kits")
    hdr = _achar_header(rows, "custo", "frete")
    if hdr is None:
        return []
    out: list[dict[str, Any]] = []
    for row in rows[hdr + 1 :]:
        sku = _sku_de(row[0] if row else "")
        if not sku.startswith("IMP-"):
            continue
        qtd = int(_f(row[1] if len(row) > 1 else 0))
        frete = round(_f(row[6] if len(row) > 6 else 0), 2)
        custo_total = round(_f(row[7] if len(row) > 7 else 0), 2)
        preco_f1 = round(_f(row[8] if len(row) > 8 else 0), 2)
        preco_f2 = round(_f(row[9] if len(row) > 9 else 0), 2)
        # col 10 = lucro F1 (pode ser R$), col 11 = val ML estoque
        lucro_f1 = round(_f(row[10] if len(row) > 10 else 0), 2)
        out.append(
            {
                "sku": sku,
                "nome": _nome_apos_sku(row[0], sku),
                "qtd_unidades": qtd,
                "tipo_cor": str(row[2] if len(row) > 2 else "").strip(),
                "custo_esmaltes": round(_f(row[3] if len(row) > 3 else 0), 2),
                "custo_complemento": round(_f(row[4] if len(row) > 4 else 0), 2),
                "custo_embalagem": round(_f(row[5] if len(row) > 5 else 0), 2),
                "frete_estimado": frete,
                "custo_total": custo_total,
                "custo_total_sem_frete": round(max(0.0, custo_total - frete), 2),
                "precos_por_fase": {
                    "fase1": preco_f1,
                    "fase2": preco_f2,
                    "fase3": round(preco_f2 * 1.08, 2) if preco_f2 else 0.0,
                },
                "preco": preco_f1,
                "lucro_ref_ml": lucro_f1,
                "margem_trabalho_pct": (
                    round(100.0 * lucro_f1 / preco_f1, 1) if preco_f1 > 0 else 0.0
                ),
                "fonte": "Consolidado_Impala_Cruzeiro.xlsx / Resumo dos Kits",
            }
        )
    return out


def parse_cores_por_kit(path: Path | None = None) -> dict[str, list[dict[str, Any]]]:
    rows = _sheet_rows(path or CONSOLIDADO_DEFAULT, "Cores por Kit")
    hdr = _achar_header(rows, "nome da cor", "código")
    if hdr is None:
        hdr = _achar_header(rows, "nome da cor", "codigo")
    if hdr is None:
        return {}
    by: dict[str, list[dict[str, Any]]] = {}
    for row in rows[hdr + 1 :]:
        sku = _sku_de(row[0] if row else "")
        if not sku.startswith("IMP-"):
            continue
        nome = str(row[1] if len(row) > 1 else "").strip()
        if not nome:
            continue
        disp = str(row[6] if len(row) > 6 else "").strip()
        by.setdefault(sku, []).append(
            {
                "nome": nome,
                "ref": str(row[2] if len(row) > 2 else "").strip(),
                "custo_un": round(_f(row[3] if len(row) > 3 else 0), 2),
                "tipo": str(row[4] if len(row) > 4 else "").strip(),
                "emb_cx": str(row[5] if len(row) > 5 else "").strip(),
                "disponivel": bool(disp) and "✗" not in disp and "nao" not in disp.lower(),
                "disponivel_txt": disp,
            }
        )
    return by


def parse_kits_cruzeiro(path: Path | None = None) -> list[dict[str, Any]]:
    rows = _sheet_rows(path or CONSOLIDADO_DEFAULT, "🎯 Kits para Validar")
    hdr = _achar_header(rows, "sku", "custo")
    if hdr is None:
        return []
    out: list[dict[str, Any]] = []
    for row in rows[hdr + 1 :]:
        sku = _sku_de(row[1] if len(row) > 1 else "")
        if not sku.startswith("CRZ-"):
            continue
        preco_ml = round(_f(row[7] if len(row) > 7 else 0), 2)
        lucro_ml = round(_f(row[8] if len(row) > 8 else 0), 2)
        margem_ml = _f(row[9] if len(row) > 9 else 0)
        if 0 < margem_ml <= 1.5:
            margem_ml *= 100.0
        preco_sh = round(_f(row[10] if len(row) > 10 else 0), 2)
        lucro_sh = round(_f(row[11] if len(row) > 11 else 0), 2)
        margem_sh = _f(row[12] if len(row) > 12 else 0)
        if 0 < margem_sh <= 1.5:
            margem_sh *= 100.0
        out.append(
            {
                "sku": sku,
                "prioridade_txt": str(row[0] or "").strip(),
                "nome": str(row[2] if len(row) > 2 else "").strip(),
                "tipo": str(row[3] if len(row) > 3 else "").strip(),
                "canal": str(row[4] if len(row) > 4 else "").strip(),
                "composicao": str(row[5] if len(row) > 5 else "").strip(),
                "custo_total": round(_f(row[6] if len(row) > 6 else 0), 2),
                "preco": preco_ml,
                "lucro_ref_ml": lucro_ml,
                "margem_trabalho_pct": round(margem_ml, 1),
                "preco_shopee": preco_sh,
                "lucro_shopee": lucro_sh,
                "margem_shopee_pct": round(margem_sh, 1),
                "linha": "cruzeiro",
                "prioridade": "P0" if str(row[0] or "").startswith("#1") else "P1",
                "fonte": "Consolidado_Impala_Cruzeiro.xlsx / Kits para Validar",
                "canais": {
                    "mercadolivre": {
                        "ativo": True,
                        "item_id": "MLB_PREENCHER",
                        "preco": preco_ml,
                        "estoque": 0,
                        "taxa_canal_pct": 14.0,
                    },
                    "shopee": {
                        "ativo": preco_sh > 0,
                        "preco": preco_sh,
                        "estoque": 0,
                    },
                },
            }
        )
    return out


def parse_ranking_margens_cruzeiro(path: Path | None = None) -> list[dict[str, Any]]:
    rows = _sheet_rows(path or CONSOLIDADO_DEFAULT, "📊 Ranking Margens")
    hdr = _achar_header(rows, "custo", "margem")
    if hdr is None:
        return []
    out: list[dict[str, Any]] = []
    for row in rows[hdr + 1 :]:
        cod = str(row[0] if row else "").strip()
        nome = str(row[1] if len(row) > 1 else "").strip()
        if not cod or not nome or not cod[0].isdigit():
            continue
        margem_ml = _f(row[6] if len(row) > 6 else 0)
        if 0 < margem_ml <= 1.5:
            margem_ml *= 100.0
        margem_sh = _f(row[8] if len(row) > 8 else 0)
        if abs(margem_sh) <= 1.5 and margem_sh != 0:
            margem_sh *= 100.0
        out.append(
            {
                "codigo": cod,
                "sku": f"CRZ-SKU-{cod}",
                "nome": nome,
                "categoria": str(row[2] if len(row) > 2 else "").strip(),
                "custo_total": round(_f(row[3] if len(row) > 3 else 0), 2),
                "preco": round(_f(row[4] if len(row) > 4 else 0), 2),
                "lucro_ref_ml": round(_f(row[5] if len(row) > 5 else 0), 2),
                "margem_trabalho_pct": round(margem_ml, 1),
                "lucro_shopee": round(_f(row[7] if len(row) > 7 else 0), 2),
                "margem_shopee_pct": round(margem_sh, 1),
                "linha": "cruzeiro_sku",
                "fonte": "Consolidado_Impala_Cruzeiro.xlsx / Ranking Margens",
            }
        )
    return out


def parse_oportunidades(path: Path | None = None) -> list[dict[str, Any]]:
    p = path or CONSOLIDADO_DEFAULT
    sheet = "Oportunidades Impala"
    rows = _sheet_rows(p, sheet)
    if not rows and INTEL_DEFAULT.is_file():
        rows = _sheet_rows(INTEL_DEFAULT, "Oportunidades Impala")
    hdr = _achar_header(rows, "estratégia") or _achar_header(rows, "estrategia")
    if hdr is None:
        hdr = _achar_header(rows, "coleção") or _achar_header(rows, "colecao")
    if hdr is None:
        return []
    out: list[dict[str, Any]] = []
    for row in rows[hdr + 1 :]:
        cor = str(row[0] if row else "").strip()
        if not cor or cor.lower().startswith("cor"):
            continue
        out.append(
            {
                "cor_produto": cor,
                "referencia": str(row[1] if len(row) > 1 else "").strip(),
                "descricao": str(row[2] if len(row) > 2 else "").strip(),
                "preco_un": round(_f(row[3] if len(row) > 3 else 0), 2),
                "colecao": str(row[4] if len(row) > 4 else "").strip(),
                "estrategia": str(row[5] if len(row) > 5 else "").strip(),
                "fonte": "Oportunidades Impala",
            }
        )
    return out


def parse_livia(path: Path | None = None) -> list[dict[str, Any]]:
    p = path or INTEL_DEFAULT
    if not p.is_file():
        p = CONSOLIDADO_DEFAULT
    rows = _sheet_rows(p, "Catálogo Lívia Distribuidora")
    hdr = _achar_header(rows, "código") or _achar_header(rows, "codigo")
    if hdr is None:
        return []
    out: list[dict[str, Any]] = []
    for row in rows[hdr + 1 :]:
        nome = str(row[0] if row else "").strip()
        if not nome or nome.lower().startswith("produto"):
            continue
        out.append(
            {
                "nome": nome,
                "codigo_livia": str(row[1] if len(row) > 1 else "").strip(),
                "descricao": str(row[2] if len(row) > 2 else "").strip(),
                "categoria": str(row[3] if len(row) > 3 else "").strip(),
                "tipo_uso": str(row[4] if len(row) > 4 else "").strip(),
                "estrategia_kit": str(row[5] if len(row) > 5 else "").strip(),
                "fonte": "Catálogo Lívia Distribuidora",
            }
        )
    return out


def parse_kits_combinados(path: Path | None = None) -> list[dict[str, Any]]:
    p = path or INTEL_DEFAULT
    if not p.is_file():
        p = CONSOLIDADO_DEFAULT
    rows = _sheet_rows(p, "Kits Combinados")
    hdr = _achar_header(rows, "composição") or _achar_header(rows, "composicao")
    if hdr is None:
        return []
    out: list[dict[str, Any]] = []
    for row in rows[hdr + 1 :]:
        nome = str(row[0] if row else "").strip()
        if not nome or nome.lower().startswith("nome"):
            continue
        out.append(
            {
                "nome": nome,
                "composicao": str(row[1] if len(row) > 1 else "").strip(),
                "custo_estimado_txt": str(row[2] if len(row) > 2 else "").strip(),
                "preco_venda_ml_txt": str(row[3] if len(row) > 3 else "").strip(),
                "lucro_liq_est_txt": str(row[4] if len(row) > 4 else "").strip(),
                "pitch": str(row[5] if len(row) > 5 else "").strip(),
                "fonte": "Kits Combinados",
            }
        )
    return out


def _produto_de_resumo(
    resumo: dict[str, Any],
    plano: dict[str, Any] | None,
    cores: list[dict[str, Any]],
) -> dict[str, Any]:
    sku = resumo["sku"]
    peso_proxy = max(280.0, float(resumo.get("qtd_unidades") or 1) * 75.0)
    invest = float((plano or {}).get("invest_validacao_reais") or 0)
    return {
        "sku": sku,
        "nome": resumo.get("nome") or sku,
        "ncm": "3304.10.00",
        "prioridade": "P0" if invest and invest <= 400 else "P1",
        "score_alavancagem": int(
            max(1, (resumo.get("margem_trabalho_pct") or 0) * (resumo.get("qtd_unidades") or 1))
        ),
        "vd_dia_ml_ref": 0,
        "vendas_historico_ml": "VALIDACAO",
        "preco_ml_mercado": resumo.get("precos_por_fase", {}).get("fase2") or resumo.get("preco"),
        "margem_trabalho_pct": resumo.get("margem_trabalho_pct") or 0,
        "lucro_ref_ml": resumo.get("lucro_ref_ml") or 0,
        "segmento": f"Validação — {resumo.get('tipo_cor') or 'Impala'}",
        "full_recomendado": True,
        "custo_unitario_esmalte": 2.87,
        "custo_esmaltes": resumo.get("custo_esmaltes") or 0,
        "custo_complemento": resumo.get("custo_complemento") or 0,
        "custo_embalagem": resumo.get("custo_embalagem") or 0,
        "custo_caixa": 0,
        "custo_total_sem_frete": resumo.get("custo_total_sem_frete") or 0,
        "frete_estimado": resumo.get("frete_estimado") or 0,
        "custo_total": resumo.get("custo_total") or 0,
        "fase_atual": 1,
        "precos_por_fase": resumo.get("precos_por_fase") or {},
        "preco": resumo.get("preco") or 0,
        "estoque_total": 0,
        "peso_gramas": int(peso_proxy),
        "invest_validacao_reais": invest,
        "valida_unidades": int((plano or {}).get("valida_unidades") or 0),
        "meta_avaliacoes": int((plano or {}).get("meta_avaliacoes") or 20),
        "acao_pos_validar": (plano or {}).get("acao_pos_validar") or "",
        "cores": [
            {
                "nome": c.get("nome"),
                "acabamento": c.get("tipo"),
                "ref": c.get("ref"),
                "disponivel": c.get("disponivel"),
            }
            for c in cores
        ],
        "fonte": resumo.get("fonte"),
        "linha": "impala_validacao",
        "canais": {
            "mercadolivre": {
                "ativo": True,
                "item_id": "MLB_PREENCHER",
                "titulo_anuncio": resumo.get("nome") or sku,
                "categoria_ml": "MLB1430",
                "preco": resumo.get("preco") or 0,
                "estoque": 0,
                "taxa_canal_pct": 18.0,
            },
            "shopee": {"ativo": False, "preco": 0, "estoque": 0},
        },
    }


def atualizar_produtos_json(
    *,
    resumos: list[dict[str, Any]],
    planos: list[dict[str, Any]],
    cores_por_sku: dict[str, list[dict[str, Any]]],
    faixas: list[dict[str, Any]],
    kits_cruzeiro: list[dict[str, Any]],
) -> dict[str, Any]:
    """Mescla validação/Cruzeiro e corrige frete por peso nos produtos existentes."""
    produtos = ler_json(CAT_PRODUTOS, default=[])
    if not isinstance(produtos, list):
        produtos = []
    by_sku = {
        str(p.get("sku") or "").upper(): p for p in produtos if isinstance(p, dict) and p.get("sku")
    }
    plano_by = {p["sku"]: p for p in planos}
    atualizados = 0
    inseridos = 0

    for p in produtos:
        if not isinstance(p, dict):
            continue
        sku = str(p.get("sku") or "").upper()
        peso = _f(p.get("peso_gramas"))
        if peso > 0 and faixas:
            novo_frete = frete_para_peso_gramas(peso, faixas)
            if novo_frete > 0:
                sem = _f(p.get("custo_total_sem_frete"))
                if sem <= 0:
                    sem = max(0.0, _f(p.get("custo_total")) - _f(p.get("frete_estimado")))
                p["frete_estimado"] = novo_frete
                p["custo_total_sem_frete"] = round(sem, 2)
                p["custo_total"] = round(sem + novo_frete, 2)
                atualizados += 1
        if sku in plano_by:
            pl = plano_by[sku]
            p["invest_validacao_reais"] = pl.get("invest_validacao_reais")
            p["valida_unidades"] = pl.get("valida_unidades")
            p["meta_avaliacoes"] = pl.get("meta_avaliacoes")
            p["acao_pos_validar"] = pl.get("acao_pos_validar")
            atualizados += 1

    for r in resumos:
        sku = r["sku"]
        if sku in by_sku:
            # só completa campos de validação/frete se faltarem
            dest = by_sku[sku]
            pl = plano_by.get(sku)
            if pl:
                dest["invest_validacao_reais"] = pl.get("invest_validacao_reais")
                dest["valida_unidades"] = pl.get("valida_unidades")
                dest["meta_avaliacoes"] = pl.get("meta_avaliacoes")
                dest["acao_pos_validar"] = pl.get("acao_pos_validar")
            if _f(dest.get("frete_estimado")) < _f(r.get("frete_estimado")):
                dest["frete_estimado"] = r.get("frete_estimado")
            continue
        novo = _produto_de_resumo(r, plano_by.get(sku), cores_por_sku.get(sku) or [])
        produtos.append(novo)
        by_sku[sku] = novo
        inseridos += 1

    for k in kits_cruzeiro:
        sku = k["sku"]
        if sku in by_sku:
            continue
        produtos.append(
            {
                **k,
                "ncm": "3304.99.90",
                "fase_atual": 1,
                "estoque_total": 0,
                "peso_gramas": 500,
                "frete_estimado": frete_para_peso_gramas(500, faixas) if faixas else 11.0,
                "vd_dia_ml_ref": 0,
                "full_recomendado": True,
                "invest_validacao_reais": round(
                    float(k.get("custo_total") or 0) * 10, 2
                ),
                "valida_unidades": 10,
                "meta_avaliacoes": 20,
            }
        )
        inseridos += 1

    escrever_json_atomico(CAT_PRODUTOS, produtos)
    return {
        "produtos_total": len(produtos),
        "produtos_atualizados": atualizados,
        "produtos_inseridos": inseridos,
    }


def emitir_metricas_planilha_ecommerce(snapshot: dict[str, Any] | None = None) -> dict[str, Any]:
    snap = snapshot or {}
    try:
        planos = snap.get("plano_validacao") or ler_json(CAT_PLANO, default={}).get("itens") or []
        cruzeiro = snap.get("kits_cruzeiro") or ler_json(CAT_CRUZEIRO, default={}).get("kits") or []
        oport = snap.get("oportunidades") or ler_json(CAT_OPORT, default={}).get("itens") or []
        livia = snap.get("livia") or ler_json(CAT_LIVIA, default={}).get("itens") or []

        invest_total = sum(float(p.get("invest_validacao_reais") or 0) for p in planos)
        gauge("catalogo.invest_validacao_total", float(invest_total))
        gauge("catalogo.plano_validacao_kits", float(len(planos)))
        for p in planos:
            from integracoes.esmaltes.metricas_catalogo_impala import kit_tag

            tags = [kit_tag(str(p.get("sku") or ""))]
            gauge("catalogo.invest_validacao", float(p.get("invest_validacao_reais") or 0), tags=tags)
            gauge("catalogo.meta_avaliacoes", float(p.get("meta_avaliacoes") or 0), tags=tags)
            gauge("catalogo.valida_unidades", float(p.get("valida_unidades") or 0), tags=tags)

        gauge("cruzeiro.kits_validacao", float(len(cruzeiro)))
        if cruzeiro:
            gauge(
                "cruzeiro.margem_media_pct",
                float(sum(float(k.get("margem_trabalho_pct") or 0) for k in cruzeiro) / len(cruzeiro)),
            )
            gauge(
                "cruzeiro.custo_investido",
                float(sum(float(k.get("custo_total") or 0) for k in cruzeiro)),
            )
            for k in cruzeiro:
                from integracoes.esmaltes.metricas_catalogo_impala import kit_tag

                tags = [kit_tag(str(k.get("sku") or "")), "linha:cruzeiro"]
                gauge("cruzeiro.preco", float(k.get("preco") or 0), tags=tags)
                gauge("cruzeiro.custo_total", float(k.get("custo_total") or 0), tags=tags)
                gauge("cruzeiro.margem_pct", float(k.get("margem_trabalho_pct") or 0), tags=tags)
                gauge("cruzeiro.lucro_ref", float(k.get("lucro_ref_ml") or 0), tags=tags)

        gauge("catalogo.oportunidades_impala", float(len(oport)))
        gauge("catalogo.complementos_livia", float(len(livia)))
        incrementar("catalogo.planilha_ecommerce_sync")
        return {
            "ok": True,
            "invest_validacao_total": round(invest_total, 2),
            "planos": len(planos),
            "cruzeiro": len(cruzeiro),
            "oportunidades": len(oport),
            "livia": len(livia),
        }
    except Exception as exc:
        logger.warning("emitir_metricas_planilha_ecommerce: %s", exc)
        incrementar("catalogo.planilha_ecommerce_sync_erro")
        return {"ok": False, "erro": str(exc)}


def sincronizar_planilhas_ecommerce(
    *,
    consolidado: Path | None = None,
    emitir_metricas: bool = True,
) -> dict[str, Any]:
    """Pipeline completo: parse → JSON catálogo → patch produtos → métricas."""
    path = consolidado or CONSOLIDADO_DEFAULT
    if not path.is_file():
        return {"ok": False, "erro": f"planilha ausente: {path}"}

    try:
        faixas = parse_frete_faixas()
        planos = parse_plano_validacao(path)
        resumos = parse_resumo_kits(path)
        cores = parse_cores_por_kit(path)
        kits_crz = parse_kits_cruzeiro(path)
        ranking = parse_ranking_margens_cruzeiro(path)
        oport = parse_oportunidades(path)
        livia = parse_livia()
        combinados = parse_kits_combinados()

        ts = datetime.now(timezone.utc).isoformat()
        escrever_json_atomico(
            CAT_FRETE,
            {"timestamp": ts, "fonte": str(EMBALAGEM_DEFAULT.name), "faixas": faixas},
        )
        escrever_json_atomico(
            CAT_PLANO,
            {
                "timestamp": ts,
                "fonte": path.name,
                "invest_total_reais": round(
                    sum(float(p.get("invest_validacao_reais") or 0) for p in planos), 2
                ),
                "itens": planos,
            },
        )
        escrever_json_atomico(
            CAT_CRUZEIRO,
            {
                "timestamp": ts,
                "fonte": path.name,
                "kits": kits_crz,
                "ranking_margens": ranking,
            },
        )
        escrever_json_atomico(
            CAT_OPORT,
            {
                "timestamp": ts,
                "itens": oport,
                "kits_combinados": combinados,
            },
        )
        escrever_json_atomico(
            CAT_LIVIA,
            {"timestamp": ts, "itens": livia},
        )

        patch = atualizar_produtos_json(
            resumos=resumos,
            planos=planos,
            cores_por_sku=cores,
            faixas=faixas,
            kits_cruzeiro=kits_crz,
        )

        snap = {
            "plano_validacao": planos,
            "kits_cruzeiro": kits_crz,
            "oportunidades": oport,
            "livia": livia,
        }
        met = emitir_metricas_planilha_ecommerce(snap) if emitir_metricas else {"ok": True, "skip": True}

        out = {
            "ok": True,
            "timestamp": ts,
            "faixas_frete": len(faixas),
            "plano_validacao": len(planos),
            "resumo_kits": len(resumos),
            "kits_cruzeiro": len(kits_crz),
            "ranking_cruzeiro": len(ranking),
            "oportunidades": len(oport),
            "livia": len(livia),
            "kits_combinados": len(combinados),
            "invest_total_reais": round(
                sum(float(p.get("invest_validacao_reais") or 0) for p in planos), 2
            ),
            "produtos": patch,
            "metricas": met,
        }
        escrever_json_atomico(ROOT / "logs" / "planilha_ecommerce_sync_ultima.json", out)
        return out
    except Exception as exc:
        logger.exception("sincronizar_planilhas_ecommerce: %s", exc)
        return {"ok": False, "erro": str(exc)}
