"""
integracoes/esmaltes/planilha_impala.py
Lê a planilha Cadastro_NCM_Bling_Impala_Cruzeiro.xlsx → cores Impala e kits cadastrados.
"""
from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any

from core.config import ROOT

logger = logging.getLogger("planilha_impala")

PLANILHA_DEFAULT = ROOT / "dados" / "Cadastro_NCM_Bling_Impala_Cruzeiro.xlsx"

_STOP = frozenset(
    {
        "esmalte",
        "impala",
        "comercial",
        "cremoso",
        "perolado",
        "metalico",
        "metálico",
        "glitter",
        "fosco",
        "acetinado",
        "transparente",
        "a",
        "cor",
        "da",
        "moda",
        "top",
        "coat",
        "base",
        "pro",
        "finish",
        "com",
        "de",
        "do",
        "e",
        "para",
        "kit",
    }
)

_PAT_COR = re.compile(
    r"(?:CREMOSO|PEROLADO|MET[AÁ]LICO|GLITTER|FOSCO|ACETINADO|TRANSPA(?:RENTE)?|"
    r"BLINDAGEM|BRILHO|ULTRA\s+SECAGEM)\s+(.+?)(?:\s+COMERCIAL)?\s*$",
    re.IGNORECASE,
)


def _norm(texto: str) -> str:
    import unicodedata

    t = unicodedata.normalize("NFKD", (texto or "").lower())
    return "".join(c for c in t if not unicodedata.combining(c))


def extrair_nome_cor(descricao: str) -> str:
    """Extrai nome comercial da cor a partir da descrição Impala."""
    desc = (descricao or "").strip()
    if not desc:
        return ""
    m = _PAT_COR.search(desc)
    if m:
        return re.sub(r"\s+", " ", m.group(1)).strip().title()
    # Fallback: remove prefixos conhecidos
    resto = desc
    for pref in (
        "ESMALTE IMPALA A COR DA MODA",
        "ESMALTE IMPALA",
        "IMPALA",
        "TOP COAT PRO FINISH IMPALA",
        "TOP COAT IMPALA",
        "IMPALA BASE",
    ):
        if resto.upper().startswith(pref):
            resto = resto[len(pref) :].strip()
    resto = re.sub(
        r"\b(CREMOSO|PEROLADO|MET[AÁ]LICO|GLITTER|FOSCO|ACETINADO|COMERCIAL)\b",
        " ",
        resto,
        flags=re.I,
    )
    resto = re.sub(r"\s+", " ", resto).strip(" -")
    return resto.title() if resto else desc.title()


def tokens_cor(nome_cor: str, descricao: str = "") -> list[str]:
    """Tokens normalizados para cruzar com títulos de kits no ML."""
    bruto = f"{nome_cor} {descricao}"
    tokens: list[str] = []
    vistos: set[str] = set()
    for parte in re.split(r"[\s/\-_]+", _norm(bruto)):
        parte = re.sub(r"[^a-z0-9]", "", parte)
        if len(parte) < 3 or parte in _STOP:
            continue
        if parte not in vistos:
            vistos.add(parte)
            tokens.append(parte)
    # nome completo da cor (útil p/ "maria cereja", "sorvete de ceu azul")
    nome_n = _norm(nome_cor)
    nome_n = re.sub(r"[^a-z0-9\s]", "", nome_n).strip()
    if len(nome_n) >= 4 and nome_n not in vistos:
        tokens.insert(0, nome_n)
    return tokens


def carregar_produtos_planilha(caminho: str | Path | None = None) -> list[dict[str, Any]]:
    """
    Lê aba 'Produtos NCM'. Retorna lista de produtos (foco Impala esmalte/top coat/base).
    Nunca lança.
    """
    path = Path(caminho) if caminho else PLANILHA_DEFAULT
    try:
        import openpyxl

        if not path.is_file():
            logger.warning("Planilha Impala ausente: %s", path)
            return []
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if "Produtos NCM" not in wb.sheetnames:
            wb.close()
            logger.warning("Aba 'Produtos NCM' ausente em %s", path)
            return []
        rows = list(wb["Produtos NCM"].iter_rows(values_only=True))
        wb.close()
        hdr_i = None
        for i, row in enumerate(rows):
            if row and str(row[0] or "").startswith("SKU"):
                hdr_i = i
                break
        if hdr_i is None:
            return []
        out: list[dict[str, Any]] = []
        for row in rows[hdr_i + 1 :]:
            if not row or not row[0]:
                continue
            sku = str(row[0]).strip()
            desc = str(row[3] or "").strip()
            marca = str(row[4] or "").strip()
            tipo = str(row[5] or "").strip()
            if not sku or not desc:
                continue
            nome_cor = extrair_nome_cor(desc)
            out.append(
                {
                    "sku": sku,
                    "referencia": str(row[1] or "").strip(),
                    "ean": str(row[2] or "").strip(),
                    "descricao": desc,
                    "marca": marca,
                    "tipo": tipo,
                    "ncm": str(row[6] or "").strip(),
                    "nome_cor": nome_cor,
                    "tokens": tokens_cor(nome_cor, desc),
                    "eh_esmalte": _norm(tipo) in ("esmalte", "top coat", "base"),
                    "eh_impala": _norm(marca) == "impala",
                }
            )
        return out
    except Exception as exc:
        logger.error("carregar_produtos_planilha erro: %s", exc)
        return []


def carregar_kits_planilha(caminho: str | Path | None = None) -> list[dict[str, Any]]:
    """Lê aba 'Kits NCM'. Nunca lança."""
    path = Path(caminho) if caminho else PLANILHA_DEFAULT
    try:
        import openpyxl

        if not path.is_file():
            return []
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if "Kits NCM" not in wb.sheetnames:
            wb.close()
            return []
        rows = list(wb["Kits NCM"].iter_rows(values_only=True))
        wb.close()
        out: list[dict[str, Any]] = []
        for row in rows:
            if not row or not isinstance(row[0], int):
                continue
            nome = str(row[1] or "").strip()
            unidades = str(row[2] or "").strip()
            m = re.search(r"(\d+)", unidades)
            qtd = int(m.group(1)) if m else 0
            out.append(
                {
                    "ordem": int(row[0]),
                    "nome": nome,
                    "unidades_txt": unidades,
                    "qtd": qtd,
                    "ncm": str(row[3] or "").strip(),
                    "observacao": str(row[5] or "").strip(),
                    "tokens": tokens_cor(nome, nome),
                }
            )
        return out
    except Exception as exc:
        logger.error("carregar_kits_planilha erro: %s", exc)
        return []


def cores_impala_disponiveis(
    produtos: list[dict[str, Any]] | None = None,
    *,
    so_esmalte: bool = True,
) -> list[dict[str, Any]]:
    """Filtra Impala (esmaltes) com nome/tokens de cor."""
    itens = produtos if produtos is not None else carregar_produtos_planilha()
    out: list[dict[str, Any]] = []
    for p in itens:
        if not p.get("eh_impala"):
            continue
        if so_esmalte and _norm(str(p.get("tipo") or "")) != "esmalte":
            continue
        if not p.get("nome_cor") and not p.get("tokens"):
            continue
        out.append(p)
    return out
