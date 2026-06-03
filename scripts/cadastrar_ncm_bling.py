#!/usr/bin/env python3
"""
scripts/cadastrar_ncm_bling.py

Cadastra/atualiza o NCM dos produtos no Bling em lote, a partir de uma planilha
SKU -> NCM que VOCÊ fornece. Habilita a emissão de NF-e (que exige NCM).

ENTRADA aceita:
  - .xlsx (aba "Produtos NCM" com colunas "SKU (código no Bling)",
    "NCM (sugerido)", "Validar c/ contador?")
  - .csv  (cabeçalho: sku,ncm[,validar])
  - .json ({"SKU": "NCM", ...})

SEGURANÇA (fiscal):
  - DRY-RUN por padrão: só mostra o que mudaria; NÃO grava.
  - --aplicar efetiva as mudanças.
  - Idempotente: se o NCM no Bling já estiver correto, pula (não regrava).
  - Itens marcados "Validar c/ contador? = Sim" são PULADOS por padrão.
    Use --incluir-validar para incluí-los (só depois de validar com o contador).

Uso:
  python scripts/cadastrar_ncm_bling.py --arquivo Cadastro_NCM_Bling_Impala_Cruzeiro.xlsx
  python scripts/cadastrar_ncm_bling.py --arquivo planilha.xlsx --aplicar
  python scripts/cadastrar_ncm_bling.py --arquivo planilha.xlsx --aplicar --incluir-validar
"""
import argparse
import csv
import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

try:
    from dotenv import load_dotenv

    load_dotenv(ROOT / ".env")
except Exception:
    pass


def _so_digitos(v) -> str:
    return "".join(ch for ch in str(v or "") if ch.isdigit())


def _ler_xlsx(caminho: Path) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(caminho, data_only=True, read_only=True)
    ws = wb["Produtos NCM"] if "Produtos NCM" in wb.sheetnames else wb.active

    # Localiza a linha de cabeçalho (tem "SKU" e "NCM")
    header_row = None
    col = {}
    for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
        textos = [str(c).strip().lower() if c is not None else "" for c in row]
        if any("sku" in t for t in textos) and any("ncm" in t for t in textos):
            header_row = ri
            for ci, t in enumerate(textos):
                if "sku" in t:
                    col["sku"] = ci
                elif "ncm" in t:
                    col["ncm"] = ci
                elif "validar" in t:
                    col["validar"] = ci
            break
    if header_row is None or "sku" not in col or "ncm" not in col:
        raise ValueError("Planilha sem cabeçalho com colunas 'SKU' e 'NCM'.")

    itens = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        sku = str(row[col["sku"]]).strip() if col["sku"] < len(row) and row[col["sku"]] is not None else ""
        ncm = _so_digitos(row[col["ncm"]]) if col["ncm"] < len(row) else ""
        validar = ""
        if "validar" in col and col["validar"] < len(row) and row[col["validar"]] is not None:
            validar = str(row[col["validar"]]).strip().lower()
        if sku and ncm:
            itens.append({"sku": sku, "ncm": ncm, "validar": validar in {"sim", "s", "true", "1"}})
    wb.close()
    return itens


def _ler_csv(caminho: Path) -> list[dict]:
    itens = []
    with caminho.open(encoding="utf-8-sig", newline="") as f:
        for linha in csv.DictReader(f):
            low = {(k or "").strip().lower(): v for k, v in linha.items()}
            sku = (low.get("sku") or "").strip()
            ncm = _so_digitos(low.get("ncm"))
            validar = (low.get("validar") or "").strip().lower() in {"sim", "s", "true", "1"}
            if sku and ncm:
                itens.append({"sku": sku, "ncm": ncm, "validar": validar})
    return itens


def _ler_json(caminho: Path) -> list[dict]:
    dados = json.loads(caminho.read_text(encoding="utf-8"))
    return [{"sku": str(k).strip(), "ncm": _so_digitos(v), "validar": False}
            for k, v in dados.items() if str(k).strip() and _so_digitos(v)]


def carregar_itens(caminho: Path) -> list[dict]:
    if not caminho.exists():
        print(f"Arquivo não encontrado: {caminho}")
        sys.exit(1)
    suf = caminho.suffix.lower()
    if suf in (".xlsx", ".xlsm"):
        return _ler_xlsx(caminho)
    if suf == ".json":
        return _ler_json(caminho)
    return _ler_csv(caminho)


def executar(itens: list[dict], aplicar: bool, incluir_validar: bool) -> dict:
    """Processa os itens. Idempotente. Devolve contadores. Não lança exceção."""
    from integracoes.bling.bling_client import buscar_produto, definir_ncm_por_sku

    res = {"total": len(itens), "atualizados": 0, "ja_corretos": 0,
           "pulados_validar": 0, "nao_encontrados": 0, "erros": 0}

    for it in itens:
        sku, ncm = it["sku"], it["ncm"]
        if it.get("validar") and not incluir_validar:
            res["pulados_validar"] += 1
            print(f"  [validar] {sku}: pulado (marque --incluir-validar após validar com o contador)")
            continue

        atual = buscar_produto(sku)
        if atual is None:
            res["nao_encontrados"] += 1
            print(f"  [não encontrado] {sku}")
            continue

        ncm_atual = _so_digitos(atual.get("ncm"))
        if ncm_atual == ncm:
            res["ja_corretos"] += 1
            continue  # idempotência: nada a fazer

        if not aplicar:
            print(f"  [dry-run] {sku}: NCM {ncm_atual or '—'} -> {ncm}")
            continue

        r = definir_ncm_por_sku(sku, ncm)
        if r.get("ok"):
            res["atualizados"] += 1
            print(f"  [ok] {sku}: NCM {ncm_atual or '—'} -> {ncm}")
        else:
            res["erros"] += 1
            print(f"  [ERRO] {sku}: {r.get('erro')}")

    return res


def main() -> int:
    parser = argparse.ArgumentParser(description="Cadastra NCM em lote no Bling.")
    parser.add_argument("--arquivo", required=True, help=".xlsx (aba Produtos NCM), .csv ou .json")
    parser.add_argument("--aplicar", action="store_true", help="Efetiva (sem isso é dry-run).")
    parser.add_argument("--incluir-validar", action="store_true",
                        help="Inclui itens marcados 'Validar c/ contador? = Sim'.")
    args = parser.parse_args()

    itens = carregar_itens(Path(args.arquivo))
    if not itens:
        print("Nenhum item válido no arquivo.")
        return 1

    modo = "APLICANDO" if args.aplicar else "DRY-RUN (nada será gravado)"
    print(f"=== Cadastro de NCM no Bling — {modo} ===")
    print(f"{len(itens)} item(ns) no arquivo.\n")

    res = executar(itens, args.aplicar, args.incluir_validar)

    print("\nResumo:")
    print(f"  total: {res['total']}")
    print(f"  {'atualizados' if args.aplicar else 'a atualizar'}: {res['atualizados'] if args.aplicar else '(ver dry-run acima)'}")
    print(f"  já corretos: {res['ja_corretos']}")
    print(f"  pulados (validar): {res['pulados_validar']}")
    print(f"  não encontrados: {res['nao_encontrados']}")
    print(f"  erros: {res['erros']}")

    # Em dry-run nunca falha; aplicando, falha só se houve erro de escrita.
    return 1 if (args.aplicar and res["erros"]) else 0


if __name__ == "__main__":
    raise SystemExit(main())
